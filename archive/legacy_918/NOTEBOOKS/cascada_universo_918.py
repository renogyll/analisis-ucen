import sys; sys.stdout.reconfigure(encoding="utf-8")
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(__file__)
XLSX = os.path.join(BASE, "tablas_pruebas_estadisticas_918.xlsx")

# ── Cargar datos ───────────────────────────────────────────────────────────────
doc  = pd.read_csv(os.path.join(BASE,"..","PROCESADO","docente_918.csv"),
                   encoding="utf-8-sig", dtype={"rut_key":str})
p3   = pd.read_csv(os.path.join(BASE,"..","PROCESADO","p3_sat_zscore_918.csv"),
                   encoding="utf-8-sig", dtype={"rut_key":str})
scat = pd.read_csv(os.path.join(BASE,"..","PROCESADO","scatter_sat_notas.csv"),
                   encoding="utf-8-sig")
edd  = pd.read_csv(os.path.join(BASE,"..","..","PROCESADO",
                                "P1_consolidado_con_evaluacion_jefes.csv"),
                   encoding="utf-8-sig", dtype={"rut_key":str})

for d in [doc, p3, edd]: d["rut_key"] = d["rut_key"].str.strip()
scat["rut_docente"] = scat["rut_docente"].astype(str).str.strip()

def tipo_p(r):
    if r["n_diplomado"]>0: return "DIPLOMADO"
    if r["n_proyecto"]>0:  return "PROYECTO"
    return "TALLER"
p3["tipo"] = p3.apply(tipo_p, axis=1)

# ── Números ────────────────────────────────────────────────────────────────────
N = 917

n_sat   = scat["rut_docente"].nunique()
n_form  = scat[scat["formado"]==True ]["rut_docente"].nunique()
n_ctrl  = scat[scat["formado"]==False]["rut_docente"].nunique()
n_nsat  = N - n_sat

n_p3    = len(p3)
n_p3_t  = (p3["tipo"]=="TALLER").sum()
n_p3_d  = (p3["tipo"]=="DIPLOMADO").sum()
n_p3_p  = (p3["tipo"]=="PROYECTO").sum()
n_no_p3 = n_form - n_p3

p3["sn"] = p3["sexo"].fillna("").str.upper().map(
    {"MUJER":"F","FEMENINO":"F","HOMBRE":"M","MASCULINO":"M","F":"F","M":"M"})
n_M  = (p3["sn"]=="M").sum()
n_F  = (p3["sn"]=="F").sum()

edd917 = edd[edd["rut_key"].isin(set(doc["rut_key"]))].copy()
edd917["formado"] = edd917["rut_key"].isin(set(p3["rut_key"]))
con_edd  = edd917[edd917["concepto"].notna()]
n_edd    = con_edd["rut_key"].nunique()
n_edd_f  = con_edd[con_edd["formado"]==True ]["rut_key"].nunique()
n_edd_c  = con_edd[con_edd["formado"]==False]["rut_key"].nunique()
n_no_edd = N - n_edd

n_sec_f = (scat["formado"]==True ).sum()
n_sec_c = (scat["formado"]==False).sum()

# ── CONSOLA ────────────────────────────────────────────────────────────────────
SEP = "═" * 68

lines = [
    "",
    SEP,
    "  UNIVERSO DE ANÁLISIS — Evaluación de Impacto Formación Docente UCEN",
    f"  {N} docentes jerarquizados con evaluación vigente",
    SEP,
    "",
    "  ┌─ PREGUNTA 1 ─────────────────────────────────────────────────────────",
    "  │  ¿Mejora el propio docente su evaluación SAT tras la formación?",
    "  │  (cambio intra-docente: z-score baseline → z-score resultado)",
    "  │",
    f"  │  {N} docentes jerarquizados",
    f"  │    └── {n_sat} con datos SAT en al menos 1 período",
    f"  │          ├── {n_form} FORMADOS (participaron en formación)",
    f"  │          │     └── {n_p3} APTOS P3 ← tienen SAT en baseline Y resultado",
    f"  │          │           ├── {n_p3_t:3d} TALLER",
    f"  │          │           ├── {n_p3_d:3d} DIPLOMADO",
    f"  │          │           └── {n_p3_p:3d} PROYECTO",
    f"  │          │           ({n_no_p3} formados sin SAT completo → excluidos)",
    f"  │          └── {n_ctrl} CONTROL (sin formación registrada)",
    f"  │          ({n_nsat} sin ningún dato SAT → excluidos de análisis SAT)",
    "  │",
    "  │  → Estadístico: Prueba t pareada · Tablas 1A (SAT nota) y 2A (% Recom.)",
    "  │    Gráficos: G1, G1.2, G2, G2.2",
    "  └──────────────────────────────────────────────────────────────────────",
    "",
    "  ┌─ PREGUNTA 2 ─────────────────────────────────────────────────────────",
    "  │  ¿Tienen los docentes formados mayor evaluación SAT que los no formados",
    "  │  en el mismo período? (comparación entre grupos por período académico)",
    "  │",
    f"  │  {n_sat} docentes con SAT",
    f"  │    ├── {n_form} FORMADOS   (6 períodos 2023-01 → 2025-02)",
    f"  │    └── {n_ctrl} CONTROL",
    "  │",
    "  │  → Estadístico: Prueba t independiente por período · Tablas 1B y 2B",
    "  │    Gráficos: G6, G6.2, G11, G11.2",
    "  └──────────────────────────────────────────────────────────────────────",
    "",
    "  ┌─ PREGUNTA 3 ─────────────────────────────────────────────────────────",
    "  │  ¿Modera la jerarquía académica el impacto de la formación en el SAT?",
    "  │",
    f"  │  {n_p3} aptos P3",
    f"  │    ├── {n_p3_t} TALLER    → 7 jerarquías con n ≥ 3",
    f"  │    ├── {n_p3_d} DIPLOMADO → 3 jerarquías con n ≥ 3",
    f"  │    └──  {n_p3_p} PROYECTO  → excluido (n insuficiente por jerarquía)",
    "  │",
    "  │  → Estadístico: ANOVA + Bonferroni · Tablas 5A–5C",
    "  │    Gráficos: G7, G7.2, G13, G13.2",
    "  └──────────────────────────────────────────────────────────────────────",
    "",
    "  ┌─ PREGUNTA 4 ─────────────────────────────────────────────────────────",
    "  │  ¿Difiere el impacto de la formación según el sexo del docente?",
    "  │",
    f"  │  {n_p3} aptos P3 → todos con sexo registrado",
    f"  │    ├── {n_M} HOMBRES",
    f"  │    └── {n_F} MUJERES",
    f"  │    (PROYECTO excluido por n insuficiente por celda sexo × tipo)",
    "  │",
    "  │  → Estadístico: Prueba t independiente · Tablas 6A (SAT) y 6B (% Recom.)",
    "  │    Gráficos: G5, G5.2",
    "  └──────────────────────────────────────────────────────────────────────",
    "",
    "  ┌─ PREGUNTA 5 ─────────────────────────────────────────────────────────",
    "  │  ¿Evalúan mejor los jefes directos a los docentes formados?",
    "  │  (EDD Total y Concepto, escala 0–1, años 2022–2025)",
    "  │",
    f"  │  {N} docentes jerarquizados",
    f"  │    └── {n_edd} con EDD registrada (al menos 1 año)",
    f"  │          ├── {n_edd_f} FORMADOS",
    f"  │          └── {n_edd_c} CONTROL",
    f"  │          ({n_no_edd} sin evaluación de jefes → excluidos)",
    "  │",
    "  │  → Estadístico: Prueba t independiente + Chi-cuadrado · Tablas 3A y 3B",
    "  │    Gráficos: G9, G10",
    "  └──────────────────────────────────────────────────────────────────────",
    "",
    "  ┌─ PREGUNTA 6 ─────────────────────────────────────────────────────────",
    "  │  ¿Aprueban más los alumnos de docentes formados?",
    "  │  (tasa aprobación ≥ 4,0 por sección, 6 períodos 2023-01 → 2025-02)",
    "  │",
    f"  │  {n_sat} docentes con calificaciones y SAT",
    f"  │    ├── {n_form} FORMADOS  → {n_sec_f:,} secciones evaluadas",
    f"  │    └── {n_ctrl} CONTROL   → {n_sec_c:,} secciones evaluadas",
    "  │",
    "  │  → Comparación de proporciones · Gráfico GN",
    "  └──────────────────────────────────────────────────────────────────────",
    "",
    SEP,
    "  Nota metodológica: 'Formado' = participó en ≥1 actividad de formación",
    "  en la BD de participacion_formacion. 'Apto P3' = formado con SAT válido",
    "  en el período inmediatamente anterior (baseline) Y posterior (resultado)",
    "  a su actividad de formación de mayor duración.",
    SEP,
    "",
]

for l in lines:
    print(l)

# ── EXCEL — insertar como hoja 0 ───────────────────────────────────────────────
wb = load_workbook(XLSX)

# Crear hoja al inicio
ws0 = wb.create_sheet("Universo de análisis", 0)
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 80

C_TITLE  = "1A237E"
C_BG     = "E8EAF6"
C_Q      = "263238"
C_Q_BG   = "ECEFF1"
C_BODY   = "212121"
C_SEP    = "90A4AE"
C_NOTE   = "546E7A"
C_GREEN  = "E8F5E9"
C_GREEN2 = "1B5E20"

def cell(ws, r, txt, bold=False, size=11, color=C_BODY, bg=None,
         italic=False, align="left", indent=0):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1, (" " * indent) + txt)
    c.font      = Font(name="Consolas", bold=bold, size=size,
                       color=color, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    ws.row_dimensions[r].height = 16
    return r + 1

r = 1
ws0.row_dimensions[r].height = 8; r += 1   # top padding

r = cell(ws0, r,
         "UNIVERSO DE ANÁLISIS — Evaluación de Impacto de la Formación Docente UCEN",
         bold=True, size=14, color="FFFFFF", bg=C_TITLE, align="center")
r = cell(ws0, r,
         f"{N} docentes jerarquizados con evaluación vigente — Universo 917",
         size=11, color=C_TITLE, bg=C_BG, align="center", italic=True)
r += 1

PREGUNTAS = [
    (
        "PREGUNTA 1 — ¿Mejora el propio docente su evaluación SAT tras la formación?",
        "Cambio intra-docente: z-score período baseline → z-score período resultado",
        [
            f"{N} docentes jerarquizados",
            f"  └── {n_sat} con datos SAT en al menos 1 período",
            f"        ├── {n_form} FORMADOS (participaron en alguna actividad de formación)",
            f"        │     └── {n_p3} APTOS P3  ←  tienen SAT en período baseline Y resultado",
            f"        │           ├── {n_p3_t:3d} tipo TALLER",
            f"        │           ├── {n_p3_d:3d} tipo DIPLOMADO",
            f"        │           └── {n_p3_p:3d} tipo PROYECTO",
            f"        │           ({n_no_p3} formados sin SAT completo → excluidos de P3)",
            f"        └── {n_ctrl} CONTROL (sin formación registrada)",
            f"        ({n_nsat} sin datos SAT → excluidos)",
        ],
        "Prueba t pareada · Tablas 1A (SAT nota) y 2A (% Recomendación) · Gráficos G1, G1.2, G2, G2.2",
    ),
    (
        "PREGUNTA 2 — ¿Tienen los docentes formados mayor SAT que los no formados?",
        "Comparación entre grupos en el mismo período académico (6 períodos 2023-01 → 2025-02)",
        [
            f"{n_sat} docentes con SAT",
            f"  ├── {n_form} FORMADOS",
            f"  └── {n_ctrl} CONTROL",
        ],
        "Prueba t independiente por período · Tablas 1B y 2B · Gráficos G6, G6.2, G11, G11.2",
    ),
    (
        "PREGUNTA 3 — ¿Modera la jerarquía académica el impacto de la formación?",
        "¿El efecto es igual para un Instructor Docente que para un Titular Regular?",
        [
            f"{n_p3} aptos P3",
            f"  ├── {n_p3_t} TALLER    →  7 jerarquías con n ≥ 3 aptos",
            f"  ├── {n_p3_d} DIPLOMADO →  3 jerarquías con n ≥ 3 aptos",
            f"  └──  {n_p3_p} PROYECTO  →  excluido (n insuficiente por celda jerarquía)",
        ],
        "ANOVA + Bonferroni · Tablas 5A–5C · Gráficos G7, G7.2, G13, G13.2",
    ),
    (
        "PREGUNTA 4 — ¿Difiere el impacto según el sexo del docente?",
        "¿El efecto de la formación es distinto para hombres que para mujeres?",
        [
            f"{n_p3} aptos P3 (todos con sexo registrado)",
            f"  ├── {n_M} HOMBRES",
            f"  └── {n_F} MUJERES",
            f"  (PROYECTO excluido: n insuficiente por celda sexo × tipo)",
        ],
        "Prueba t independiente · Tablas 6A (SAT nota) y 6B (% Recom.) · Gráficos G5, G5.2",
    ),
    (
        "PREGUNTA 5 — ¿Evalúan mejor los jefes a los docentes formados?",
        "Evaluación de Desempeño Docente (EDD) por directores/jefes, escala 0–1, años 2022–2025",
        [
            f"{N} docentes jerarquizados",
            f"  └── {n_edd} con EDD registrada (al menos 1 año evaluado)",
            f"        ├── {n_edd_f} FORMADOS",
            f"        └── {n_edd_c} CONTROL",
            f"        ({n_no_edd} sin evaluación de jefes → excluidos de este análisis)",
        ],
        "Prueba t independiente + Chi-cuadrado · Tablas 3A y 3B · Gráficos G9, G10",
    ),
    (
        "PREGUNTA 6 — ¿Aprueban más los alumnos de docentes formados?",
        "Tasa de aprobación (nota ≥ 4,0) por sección, 6 períodos 2023-01 → 2025-02",
        [
            f"{n_sat} docentes con calificaciones y SAT registrados",
            f"  ├── {n_form} FORMADOS  →  {n_sec_f:,} secciones evaluadas",
            f"  └── {n_ctrl} CONTROL   →  {n_sec_c:,} secciones evaluadas",
        ],
        "Comparación de proporciones · Gráfico GN_aprobacion",
    ),
]

Q_COLORS = ["1565C0","6A1B9A","00695C","E65100","B71C1C","2E7D32"]
Q_BG     = ["E3F2FD","F3E5F5","E0F2F1","FFF3E0","FFEBEE","E8F5E9"]

for qi, (titulo, subtitulo, cuerpo, estadistico) in enumerate(PREGUNTAS):
    qc   = Q_COLORS[qi]
    qbg  = Q_BG[qi]

    r = cell(ws0, r, f"  ▌  {titulo}",
             bold=True, size=11, color=qc, bg=qbg)
    r = cell(ws0, r, f"     {subtitulo}",
             italic=True, size=10, color=C_NOTE, bg=qbg)
    ws0.row_dimensions[r-1].height = 14
    r = cell(ws0, r, "", bg="FFFFFF")  # línea vacía

    for linea in cuerpo:
        r = cell(ws0, r, f"       {linea}", size=10, color=C_BODY)

    r = cell(ws0, r, "", bg="FFFFFF")
    r = cell(ws0, r, f"     → {estadistico}",
             italic=True, size=9, color=C_NOTE)
    r += 1

r = cell(ws0, r,
         "Nota: 'Formado' = participó en ≥1 actividad en la BD participacion_formacion. "
         "'Apto P3' = formado con SAT válido en período baseline Y resultado.",
         italic=True, size=9, color=C_NOTE, bg=C_BG)
r = cell(ws0, r,
         "Las poblaciones no son mutuamente excluyentes entre preguntas: "
         "un mismo docente puede aparecer en múltiples análisis.",
         italic=True, size=9, color=C_NOTE, bg=C_BG)

wb.save(XLSX)
print(f"\nHoja 'Universo de análisis' insertada en posición 0 de {XLSX}")
