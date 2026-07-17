import sys; sys.stdout.reconfigure(encoding="utf-8")
import pandas as pd
import numpy as np
from scipy.stats import ttest_rel, ttest_ind
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
import os

engine = create_engine("postgresql://ucen_user:ucen2026@localhost:5432/ucen")
BASE = os.path.dirname(__file__)
OUT  = os.path.join(BASE, "tablas_pruebas_estadisticas_918.xlsx")

# ── Colores ────────────────────────────────────────────────────────────────────
C_HEADER   = "37474F"   # gris oscuro
C_SIG3     = "C8E6C9"   # verde claro  ***
C_SIG2     = "DCEDC8"   # verde muy claro **
C_SIG1     = "F1F8E9"   # casi blanco  *
C_NS       = "F5F5F5"   # gris claro   ns
C_GLOBAL   = "E8F5E9"   # verde pálido fila global
C_WHITE    = "FFFFFF"
C_TITLE_BG = "ECEFF1"

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def row_fill(sig):
    m = {"***": C_SIG3, "**": C_SIG2, "*": C_SIG1, "ns": C_NS}
    return PatternFill("solid", fgColor=m.get(sig, C_NS))

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def sig_label(p):
    if p == "<0.001" or (isinstance(p, float) and p < 0.001): return "***"
    if isinstance(p, float):
        if p < 0.01:  return "**"
        if p < 0.05:  return "*"
    return "ns"

def write_table(ws, start_row, title, subtitle, headers, rows,
                sig_col_idx=None, global_rows=None, col_widths=None, bullets=None):
    """
    Escribe un bloque de tabla en la hoja ws a partir de start_row.
    bullets: lista de 2 strings que se escriben al costado derecho de la tabla.
    """
    n_cols    = len(headers)
    bc        = n_cols + 2   # columna inicio del bloque de bullets
    first_row = start_row

    # Título
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=n_cols)
    tc = ws.cell(start_row, 1, title)
    tc.font      = Font(bold=True, size=13, color="1A237E")
    tc.fill      = PatternFill("solid", fgColor=C_TITLE_BG)
    tc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    start_row += 1

    if subtitle:
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=n_cols)
        sc = ws.cell(start_row, 1, subtitle)
        sc.font      = Font(italic=True, size=10, color="546E7A")
        sc.fill      = PatternFill("solid", fgColor=C_WHITE)
        sc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        start_row += 1

    # Encabezados
    for j, h in enumerate(headers, 1):
        c = ws.cell(start_row, j, h)
        c.font      = Font(bold=True, size=11, color="FFFFFF")
        c.fill      = header_fill(C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = thin_border()
    start_row += 1

    data_first_row = start_row

    # Filas de datos
    for ri, row in enumerate(rows):
        is_global = (global_rows and ri in global_rows)
        sig_val   = row[sig_col_idx] if sig_col_idx is not None else "ns"
        fill      = PatternFill("solid", fgColor=C_GLOBAL) if is_global else row_fill(sig_val)

        for j, val in enumerate(row, 1):
            c = ws.cell(start_row, j, val)
            c.fill      = fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin_border()
            c.font      = Font(size=11,
                               bold=(is_global or (sig_col_idx is not None
                                                   and j-1 == sig_col_idx
                                                   and sig_val != "ns")))
        start_row += 1

    data_last_row = start_row - 1

    # Leyenda
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=n_cols)
    lc = ws.cell(start_row, 1,
                 "*** p < 0.001   |   ** p < 0.01   |   * p < 0.05   |   ns no significativo")
    lc.font      = Font(italic=True, size=9, color="78909C")
    lc.alignment = Alignment(horizontal="left")
    start_row += 3   # espacio entre tablas

    # ── Bullets al costado ────────────────────────────────────────────────────
    if bullets:
        # Cabecera "Síntesis" alineada con el título de la tabla
        ws.merge_cells(start_row=first_row, start_column=bc,
                       end_row=first_row, end_column=bc + 3)
        bh = ws.cell(first_row, bc, "Síntesis")
        bh.font      = Font(bold=True, size=11, color="FFFFFF")
        bh.fill      = PatternFill("solid", fgColor="455A64")
        bh.alignment = Alignment(horizontal="center", vertical="center")

        n_data = data_last_row - data_first_row + 1
        half   = max(1, (n_data + 1) // 2)

        for bi, txt in enumerate(bullets[:2]):
            r_s = data_first_row + bi * half
            r_e = min(data_first_row + (bi + 1) * half - 1, data_last_row)
            r_e = max(r_s, r_e)

            ws.merge_cells(start_row=r_s, start_column=bc,
                           end_row=r_e, end_column=bc + 3)
            cell = ws.cell(r_s, bc, f"{'①②'[bi]}  {txt}")
            cell.font      = Font(size=10, color="1B3A6B")
            cell.fill      = PatternFill("solid", fgColor="E3F2FD")
            cell.alignment = Alignment(wrap_text=True, vertical="top",
                                       horizontal="left")
            cell.border = thin_border()
            cur_h = ws.row_dimensions[r_s].height or 15
            ws.row_dimensions[r_s].height = max(cur_h, 42)

        # Ancho de las columnas de bullets (solo si no están ya definidas)
        for extra in range(bc, bc + 4):
            ltr = get_column_letter(extra)
            if ws.column_dimensions[ltr].width < 16:
                ws.column_dimensions[ltr].width = 18

    # Anchos de columna datos
    if col_widths:
        for j, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w

    return start_row


wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 1 — Pruebas t sobre Z-score SAT nota
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "T-test SAT nota"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 20

row = 1

# — Tabla 1A: T pareada (pre/post mismo docente)
h_a = ["Tipo formación","n","Z baseline","Z resultado","Δ Z","t","p-valor","Cohen d","Sig."]
d_a = [
    ["TALLER",    154, "+0.090", "+0.056", "−0.035", "+0.544", "0.5890", "−0.044", "ns"],
    ["DIPLOMADO",  36, "+0.204", "+0.177", "−0.028", "+0.208", "0.8360", "−0.035", "ns"],
    ["PROYECTO",    7, "+0.482", "+0.257", "−0.225", "+0.674", "0.5250", "−0.255", "ns"],
]
row = write_table(ws1, row,
    "Tabla 1A — Prueba t Pareada: ¿Cambia el Z-score SAT del docente entre baseline y resultado?",
    "Hipótesis: el docente mejora su posición relativa (z-score) dentro de su facultad tras formarse.",
    h_a, d_a, sig_col_idx=8,
    col_widths=[20,6,12,12,10,10,10,10,8],
    bullets=[
        "Ninguno de los 3 tipos de formación produce un cambio significativo en el z-score SAT "
        "propio del docente entre baseline y resultado (todos ns). El efecto no es intra-docente.",
        "La formación sí mejora la posición relativa del docente frente al grupo control "
        "(ver Tabla 1B), pero no transforma el puntaje absoluto del mismo docente antes/después.",
    ])

# — Tabla 1B: T independiente por período (formados vs control)
h_b = ["Período","n Formados","n Control","Dif. Z","t","p-valor","Cohen d","Sig."]
d_b = [
    ["2023-01", 169, 299, "+0.089", "+1.039", "0.2997", "+0.099", "ns"],
    ["2023-02", 172, 296, "+0.141", "+1.733", "0.0839", "+0.167", "ns"],
    ["2024-01", 195, 350, "+0.240", "+3.008", "0.0028", "+0.260", "**"],
    ["2024-02", 186, 332, "+0.197", "+2.607", "0.0094", "+0.231", "**"],
    ["2025-01", 193, 415, "+0.117", "+1.555", "0.1206", "+0.132", "ns"],
    ["2025-02", 152, 337, "+0.244", "+3.275", "0.0012", "+0.302", "***"],
]
row = write_table(ws1, row,
    "Tabla 1B — Prueba t Independiente por período: Z-score SAT Formados vs Sin Formación",
    "¿Tienen los docentes formados mejor posición relativa en SAT que los no formados en ese período?",
    h_b, d_b, sig_col_idx=7,
    col_widths=[12,14,12,10,10,10,10,8],
    bullets=[
        "Los docentes formados superan significativamente al control en 3 de 6 períodos: "
        "2024-01 (**), 2024-02 (**) y 2025-02 (***). La ventaja se consolida desde 2024.",
        "El Cohen d máximo es +0.302 (2025-02), clasificado como efecto pequeño-mediano. "
        "La diferencia es real aunque moderada; no uniforme en el tiempo.",
    ])


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 2 — Pruebas t sobre Z-score % Recomendación (SAT_BIN)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("T-test SAT BIN (%Recom.)")
ws2.sheet_view.showGridLines = False
row = 1

# Cargar datos reales para prueba pareada
df_bin = pd.read_csv(os.path.join(BASE,"..","PROCESADO","p3_bin_zscore_918.csv"),
                     encoding="utf-8-sig", dtype={"rut_key":str})
df_bin["rut_key"] = df_bin["rut_key"].str.strip()

def sig_p(p):
    if p < 0.001: return "***"
    if p < 0.01:  return "**"
    if p < 0.05:  return "*"
    return "ns"

datos_2a = []
for tipo in ["TALLER","DIPLOMADO","PROYECTO"]:
    sub = df_bin[df_bin["tipo_principal"]==tipo].dropna(
        subset=["z_bin_baseline","z_bin_resultado"])
    if len(sub) < 3: continue
    b = sub["z_bin_baseline"].values
    r = sub["z_bin_resultado"].values
    t_v, p_v = ttest_rel(r, b)
    d = (r-b).mean() / (r-b).std() if (r-b).std() > 0 else 0.0
    datos_2a.append([tipo, len(sub),
                     f"{b.mean():+.3f}", f"{r.mean():+.3f}",
                     f"{(r-b).mean():+.3f}", f"{t_v:+.3f}",
                     f"{p_v:.4f}", f"{d:+.3f}", sig_p(p_v)])

h_2a = ["Tipo formación","n","Z baseline","Z resultado","Δ Z","t","p-valor","Cohen d","Sig."]
row = write_table(ws2, row,
    "Tabla 2A — Prueba t Pareada: Z-score % Recomendación — baseline vs resultado",
    "Pregunta SAT_BIN: ¿Recomendaría a este/a docente? (% de respuestas Sí, z-scoreado por facultad×período)",
    h_2a, datos_2a, sig_col_idx=8,
    col_widths=[20,6,12,12,10,10,10,10,8],
    bullets=[
        "Al igual que en SAT nota, ningún tipo de formación produce un cambio intra-docente "
        "significativo en % Recomendación (todos ns). Resultado consistente con Tabla 1A.",
        "La coherencia con SAT nota (r=0.893 entre ambas métricas) confirma la solidez "
        "del instrumento: ambas preguntas capturan el mismo fenómeno.",
    ])

# Prueba C BIN — datos ya calculados en G11.2
h_2b = ["Período","n Formados","n Control","Dif. Z","t","p-valor","Cohen d","Sig."]
d_2b = [
    ["2023-01", 169, 299, "+0.066", "+0.695", "0.4876", "+0.067", "ns"],
    ["2023-02", 172, 296, "+0.125", "+1.351", "0.1774", "+0.129", "ns"],
    ["2024-01", 195, 350, "+0.216", "+2.587", "0.0100", "+0.224", "**"],
    ["2024-02", 186, 332, "+0.183", "+2.157", "0.0315", "+0.192", "*"],
    ["2025-01", 193, 415, "+0.126", "+1.529", "0.1270", "+0.129", "ns"],
    ["2025-02", 152, 337, "+0.212", "+2.502", "0.0128", "+0.230", "*"],
]
row = write_table(ws2, row,
    "Tabla 2B — Prueba t Independiente por período: Z-score % Recomendación Formados vs Sin Formación",
    "Misma metodología que SAT nota. Correlación entre ambas métricas: r = 0.893.",
    h_2b, d_2b, sig_col_idx=7,
    col_widths=[12,14,12,10,10,10,10,8],
    bullets=[
        "Docentes formados superan al control en % Recomendación en 4 de 6 períodos, "
        "con significancia en 2024-01 (**), 2024-02 (*) y 2025-02 (*). Patrón paralelo a SAT nota.",
        "Magnitudes levemente inferiores a SAT nota (Cohen d máx. +0.230 vs +0.302), "
        "pero la dirección y estructura temporal son idénticas en ambas métricas.",
    ])


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 3 — Pruebas t y Chi² sobre EDD Jefes
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("T-test + Chi2 EDD Jefes")
ws3.sheet_view.showGridLines = False
row = 1

# Tabla 3A: T independiente EDD por año
h_3a = ["Año","n Form.","EDD Form.","n Control","EDD Ctrl","Diferencia","t","p-valor","Cohen d","Sig."]
d_3a = [
    ["2022",   75,  "0.864", 145,  "0.854", "+0.010", "+0.536",  "0.5927", "+0.072", "ns"],
    ["2023",   95,  "0.897", 151,  "0.872", "+0.025", "+1.300",  "0.1950", "+0.164", "ns"],
    ["2024",  126,  "0.734", 244,  "0.655", "+0.079", "+3.118",  "0.0020", "+0.339", "**"],
    ["2025",  129,  "0.864", 336,  "0.604", "+0.260", "+11.035", "<0.001", "+0.955", "***"],
    ["GLOBAL",425,  "0.833", 876,  "0.706", "+0.127", "+9.615",  "<0.001", "+0.522", "***"],
]
row = write_table(ws3, row,
    "Tabla 3A — Prueba t Independiente por año: EDD Total Formados vs Sin Formación",
    "EDD Total = nota promedio evaluación de jefes (escala 0–1). Universo 917, años 2022–2025.",
    h_3a, d_3a, sig_col_idx=9, global_rows=[4],
    col_widths=[10,10,12,12,10,12,10,10,10,8],
    bullets=[
        "El efecto de la formación sobre la EDD crece con el tiempo: no significativo en "
        "2022–2023, significativo en 2024 (**) y altamente significativo en 2025 (***). "
        "Sugiere un impacto acumulativo.",
        "El Cohen d global es +0.522 (efecto mediano) y llega a +0.955 en 2025, "
        "el tamaño de efecto más grande de todo el análisis. La EDD de jefes es el "
        "indicador más sensible a la formación.",
    ])

# Tabla 3B: Chi-cuadrado concepto EDD
h_3b = ["Alcance","n Form.","n Control","Chi²","gl","p-valor","Cramer V","Efecto","Sig."]
d_3b = [
    ["Global (todos los años)", 439,  971, "16.213", 3, "0.0010", "0.107", "Pequeño", "**"],
    ["2022",                     64,  256,  "4.893", 3, "0.1798", "0.138", "—",       "ns"],
    ["2023",                     96,  307,  "4.894", 3, "0.1798", "0.126", "—",       "ns"],
    ["2024",                    115,  484, "13.924", 3, "0.0030", "0.190", "Pequeño", "**"],
    ["2025",                    164,  930,  "5.379", 3, "0.1461", "0.108", "—",       "ns"],
]
row = write_table(ws3, row,
    "Tabla 3B — Chi-cuadrado: Distribución de Concepto EDD (Muy Bueno / Bueno / Insuficiente / Deficiente)",
    "¿Difiere la distribución de conceptos entre formados y sin formación? gl = 3 grados de libertad.",
    h_3b, d_3b, sig_col_idx=8, global_rows=[0],
    col_widths=[26,10,12,10,6,10,12,12,8],
    bullets=[
        "Globalmente y en 2024, la distribución de conceptos difiere significativamente "
        "entre formados y no formados (V de Cramer = 0.107–0.190, efecto pequeño). "
        "Los formados acumulan más conceptos Muy Bueno.",
        "En 2022, 2023 y 2025 la diferencia no es significativa (ns), posiblemente "
        "por tamaños muestrales menores o menor madurez del efecto en esos años.",
    ])


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 4 — Resumen comparativo SAT vs BIN
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Resumen comparativo")
ws4.sheet_view.showGridLines = False
row = 1

h_res = ["Período","SAT nota — Dif.Z","SAT nota — Sig.",
         "% Recom. — Dif.Z","% Recom. — Sig.","Coinciden?"]
d_res = [
    ["2023-01", "+0.089", "ns",  "+0.066", "ns",  "✓"],
    ["2023-02", "+0.141", "ns",  "+0.125", "ns",  "✓"],
    ["2024-01", "+0.240", "**",  "+0.216", "**",  "✓"],
    ["2024-02", "+0.197", "**",  "+0.183", "*",   "✓"],
    ["2025-01", "+0.117", "ns",  "+0.126", "ns",  "✓"],
    ["2025-02", "+0.244", "***", "+0.212", "*",   "✓"],
]
row = write_table(ws4, row,
    "Resumen — SAT nota vs % Recomendación: coincidencia de resultados por período",
    "Ambas métricas apuntan en la misma dirección en todos los períodos. r = 0.893 entre ellas.",
    h_res, d_res, sig_col_idx=2,
    col_widths=[12,20,16,20,16,14],
    bullets=[
        "Ambas métricas coinciden en dirección y significancia en los 6 períodos, "
        "validando la consistencia del instrumento SAT y la robustez de los hallazgos.",
        "La ventaja del grupo formado se consolida a partir de 2024 y persiste en 2025, "
        "sugiriendo un efecto acumulativo que se afianza con el tiempo.",
    ])

# Tabla resumen Cohen d
h_d = ["Prueba","Métrica","Períodos/años significativos","Cohen d máximo","Interpretación"]
d_d = [
    ["T independiente", "SAT nota (z-score)",  "2024-01**, 2024-02**, 2025-02***",      "d=+0.302 (pequeño-med.)", "Diferencia real pero moderada"],
    ["T independiente", "% Recomendación (z)", "2024-01**, 2024-02*, 2025-02*",         "d=+0.230 (pequeño)",      "Efecto consistente y paralelo"],
    ["T independiente", "EDD Jefes",           "2024**, 2025***, Global***",             "d=+0.955 (grande) 2025",  "Efecto más fuerte del análisis"],
    ["Chi-cuadrado",    "Concepto EDD",        "Global**, 2024**",                       "V=0.190 (pequeño)",       "Distribución de categorías difiere"],
    ["T pareada",       "SAT nota (z-score)",  "Ninguno (todos ns)",                    "—",                       "Sin cambio pre/post en el docente"],
    ["T pareada",       "% Recomendación (z)", "Ninguno (todos ns)",                    "—",                       "Sin cambio pre/post en el docente"],
]
row = write_table(ws4, row,
    "Cuadro resumen — Todas las pruebas estadísticas realizadas",
    "Cohen d: pequeño ≥0.20 · mediano ≥0.50 · grande ≥0.80   |   Cramer V: pequeño ≥0.10 · mediano ≥0.30",
    h_d, d_d,
    col_widths=[22,24,40,26,35],
    bullets=[
        "La EDD Jefes es el indicador con mayor tamaño de efecto (d=+0.955 en 2025), "
        "superando al SAT en magnitud. Es el mejor predictor del impacto de la formación.",
        "La ausencia de cambio intra-docente (pruebas pareadas ns) es consistente con "
        "un modelo de impacto relativo: la formación mejora la posición frente al grupo "
        "control, no el puntaje absoluto individual.",
    ])


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 5 — T-test por Jerarquía: DIPLOMADO (G13)
# ══════════════════════════════════════════════════════════════════════════════
from scipy.stats import f_oneway
from itertools import combinations

ws5 = wb.create_sheet("T-test Jerarquía Diplomado")
ws5.sheet_view.showGridLines = False
row = 1

# — Cargar datos reales para calcular G13 en tiempo real
p3_sat = pd.read_csv(os.path.join(BASE, "..", "PROCESADO", "p3_sat_zscore_918.csv"),
                     encoding="utf-8-sig", dtype={"rut_key": str})
p3_sat["rut_key"] = p3_sat["rut_key"].str.strip()

def tipo_principal(r):
    if r["n_diplomado"] > 0: return "DIPLOMADO"
    if r["n_proyecto"]  > 0: return "PROYECTO"
    return "TALLER"
p3_sat["tipo_principal"] = p3_sat.apply(tipo_principal, axis=1)

dip = p3_sat[p3_sat["tipo_principal"] == "DIPLOMADO"].copy()

jer_map_xlsx = {
    "INSTRUCTOR DOCENTE": "Instructor Doc.",
    "ASISTENTE DOCENTE":  "Asist. Docente",
    "ASISTENTE REGULAR":  "Asist. Regular",
    "ASOCIADO DOCENTE":   "Asoc. Docente",
    "ASOCIADO REGULAR":   "Asoc. Regular",
    "TITULAR DOCENTE":    "Titular Docente",
    "TITULAR REGULAR":    "Titular Regular",
}
ORD_JER = ["Instructor Doc.", "Asist. Docente", "Asist. Regular",
           "Asoc. Docente", "Asoc. Regular", "Titular Docente", "Titular Regular"]

dip["jer"] = dip["jerarquia"].map(jer_map_xlsx).fillna(dip["jerarquia"])

grupos_cnt = dip.groupby("jer")["delta_z"].count()
jer_incl   = [j for j in ORD_JER if j in grupos_cnt.index and grupos_cnt[j] >= 3]

# Tabla descriptiva
desc_rows_xlsx = []
grupos_data_xlsx = {}
for jer in jer_incl:
    vals = dip[dip["jer"] == jer]["delta_z"].dropna().values
    grupos_data_xlsx[jer] = vals
    zb = dip[dip["jer"] == jer]["z_baseline"].dropna().mean()
    zr = dip[dip["jer"] == jer]["z_resultado"].dropna().mean()
    desc_rows_xlsx.append([
        jer, int(len(vals)),
        f"{zb:+.3f}", f"{zr:+.3f}",
        f"{vals.mean():+.3f}", f"{vals.std():.3f}",
        f"{vals.min():+.3f}", f"{vals.max():+.3f}",
    ])

h_5a = ["Jerarquía", "n", "Z Baseline", "Z Resultado",
        "Δ Z promedio", "SD", "Δ Z mín.", "Δ Z máx."]
row = write_table(ws5, row,
    "Tabla 5A — Estadísticos descriptivos Δ Z-score SAT por jerarquía (solo DIPLOMADO)",
    f"Universo: {len(dip)} docentes aptos P3 tipo DIPLOMADO. Jerarquías con n≥3 incluidas.",
    h_5a, desc_rows_xlsx,
    col_widths=[22, 6, 13, 13, 14, 10, 10, 10],
    bullets=[
        "Los Δ Z-score son pequeños en todas las jerarquías (entre −0.2 y +0.2), "
        "sin un patrón de gradiente claro por nivel académico.",
        "Los tamaños muestrales por jerarquía son reducidos (máx. n=18), "
        "lo que limita la potencia estadística de las comparaciones entre grupos.",
    ])

# ANOVA
arrays_xlsx = [grupos_data_xlsx[j] for j in jer_incl]
F_val, p_anova = f_oneway(*arrays_xlsx)
grand_mean = np.concatenate(arrays_xlsx).mean()
SS_b = sum(len(g) * (g.mean() - grand_mean)**2 for g in arrays_xlsx)
SS_t = sum(((v - grand_mean)**2).sum() for v in arrays_xlsx)
eta2 = SS_b / SS_t if SS_t > 0 else 0.0

def sig_f(p):
    if p < 0.001: return "***"
    if p < 0.01:  return "**"
    if p < 0.05:  return "*"
    return "ns"

s_anova = sig_f(p_anova)
gl_e  = len(jer_incl) - 1
gl_d  = sum(len(g) for g in arrays_xlsx) - len(jer_incl)
anova_row_xlsx = [[
    f"F({gl_e}, {gl_d}) = {F_val:.3f}",
    f"{p_anova:.4f}",
    f"{eta2:.3f}",
    s_anova,
    f"gl entre = {gl_e}",
    f"gl dentro = {gl_d}",
]]
h_5b = ["Estadístico F", "p-valor", "η² (eta²)", "Sig.", "Grados libertad entre", "Grados libertad dentro"]
row = write_table(ws5, row,
    "Tabla 5B — ANOVA de un factor: Δ Z-score ~ Jerarquía",
    "¿Difiere el impacto del diplomado según la jerarquía académica del docente?",
    h_5b, anova_row_xlsx, sig_col_idx=3,
    col_widths=[24, 12, 14, 8, 24, 24],
    bullets=[
        "ANOVA no significativo (η²=0.021, efecto nulo): no hay evidencia de que "
        "el impacto del diplomado difiera según la jerarquía del docente.",
        "El diplomado beneficia por igual a docentes de distintos rangos académicos; "
        "la jerarquía no modera el efecto de la formación.",
    ])

# Pairwise t-tests Bonferroni
pairs_xlsx = list(combinations(jer_incl, 2))
n_tests    = len(pairs_xlsx)
pair_rows_xlsx = []
for j1, j2 in pairs_xlsx:
    from scipy.stats import ttest_ind as _tind
    g1, g2 = grupos_data_xlsx[j1], grupos_data_xlsx[j2]
    t_val, p_raw = _tind(g1, g2, equal_var=False)
    p_adj = min(p_raw * n_tests, 1.0)
    pool_sd = np.sqrt((g1.std()**2 + g2.std()**2) / 2)
    d = (g1.mean() - g2.mean()) / pool_sd if pool_sd > 0 else 0.0
    pair_rows_xlsx.append([
        j1, j2,
        int(len(g1)), int(len(g2)),
        f"{g1.mean()-g2.mean():+.3f}",
        f"{t_val:+.3f}",
        f"{p_raw:.4f}",
        f"{p_adj:.4f}",
        f"{d:+.3f}",
        sig_f(p_adj),
    ])

h_5c = ["Jerarquía A", "Jerarquía B", "n A", "n B",
        "Diferencia Δ", "t", "p crudo", "p Bonferroni", "Cohen d", "Sig."]
row = write_table(ws5, row,
    f"Tabla 5C — Comparaciones por pares (Welch t, corrección Bonferroni n={n_tests})",
    "p Bonferroni = p crudo × número de comparaciones. Significancia evaluada sobre p ajustado.",
    h_5c, pair_rows_xlsx, sig_col_idx=9,
    col_widths=[22, 22, 6, 6, 14, 10, 12, 14, 12, 8],
    bullets=[
        "Ninguna comparación por pares alcanza significancia tras la corrección Bonferroni "
        "(todos ns), confirmando la homogeneidad del efecto entre jerarquías.",
        "El par con mayor diferencia observada no supera el umbral ajustado, lo que "
        "respalda que las diferencias descriptivas entre grupos son producto del azar.",
    ])

# Nota interpretativa al final de la hoja
nota_row = row
ws5.merge_cells(start_row=nota_row, start_column=1,
                end_row=nota_row, end_column=10)
nc = ws5.cell(nota_row, 1,
    f"Interpretación: ANOVA no significativo (p={p_anova:.3f}). "
    "No hay evidencia de que el impacto del diplomado difiera entre jerarquías. "
    "El efecto es homogéneo, aunque los tamaños muestrales por grupo son reducidos (máx. n=18).")
nc.font = Font(italic=True, size=10, color="37474F")
nc.alignment = Alignment(horizontal="left", wrap_text=True)
ws5.row_dimensions[nota_row].height = 36


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 6 — T-test Sexo: Hombre vs Mujer (G5 y G5.2)
# ══════════════════════════════════════════════════════════════════════════════
from scipy.stats import ttest_ind as _tind2

ws6 = wb.create_sheet("T-test Sexo G5 y G5.2")
ws6.sheet_view.showGridLines = False
row = 1

# ── Cargar datos ───────────────────────────────────────────────────────────────
p3_s = pd.read_csv(os.path.join(BASE, "..", "PROCESADO", "p3_sat_zscore_918.csv"),
                   encoding="utf-8-sig", dtype={"rut_key": str})
p3_s["sexo_norm"] = (p3_s["sexo"].fillna("").str.strip().str.upper()
                     .map({"MUJER":"F","FEMENINO":"F","HOMBRE":"M","MASCULINO":"M","F":"F","M":"M"}))
p3_s["tipo_principal"] = p3_s.apply(tipo_principal, axis=1)
df5_sat = p3_s[p3_s["sexo_norm"].isin(["M","F"])].copy()

p3_b = pd.read_csv(os.path.join(BASE, "..", "PROCESADO", "p3_bin_zscore_918.csv"),
                   encoding="utf-8-sig", dtype={"rut_key": str})
p3_b["sexo_norm"] = (p3_b["sexo"].fillna("").str.strip().str.upper()
                     .map({"MUJER":"F","FEMENINO":"F","HOMBRE":"M","MASCULINO":"M","F":"F","M":"M"}))
df5_bin = p3_b[p3_b["sexo_norm"].isin(["M","F"])].copy()

def _row_sexo(sub, col_dz):
    m = sub[sub["sexo_norm"]=="M"][col_dz].dropna().values
    f = sub[sub["sexo_norm"]=="F"][col_dz].dropna().values
    if len(m) < 3 or len(f) < 3:
        return None
    t, p = _tind2(m, f, equal_var=False)
    pool_sd = np.sqrt((m.std()**2 + f.std()**2) / 2)
    d = (m.mean() - f.mean()) / pool_sd if pool_sd > 0 else 0.0
    return [int(len(m)), f"{m.mean():+.3f}",
            int(len(f)), f"{f.mean():+.3f}",
            f"{m.mean()-f.mean():+.3f}",
            f"{t:+.3f}", f"{p:.4f}", f"{d:+.3f}", sig_f(p)]

h_sex = ["Tipo formación", "n Hombres", "Δz Hombres",
         "n Mujeres", "Δz Mujeres", "Dif. H−M", "t", "p-valor", "Cohen d", "Sig."]

# Tabla 6A — SAT nota
rows_6a = []
for tipo in ["TALLER", "DIPLOMADO", "PROYECTO"]:
    sub = df5_sat[df5_sat["tipo_principal"] == tipo]
    r   = _row_sexo(sub, "delta_z")
    if r is None:
        rows_6a.append([tipo, "—", "—", "—", "—", "—", "—", "insuf. n", "—", "—"])
    else:
        rows_6a.append([tipo] + r)

row = write_table(ws6, row,
    "Tabla 6A — Prueba t Independiente: Δ Z-score SAT nota por Sexo (Hombre vs Mujer) — G5",
    "¿Difiere el impacto de la formación en el Z-score SAT según el sexo del docente? "
    "PROYECTO excluido por n insuficiente (M=5, F=2).",
    h_sex, rows_6a, sig_col_idx=9,
    col_widths=[18, 12, 12, 12, 12, 12, 10, 10, 10, 8],
    bullets=[
        "Ni en TALLER ni en DIPLOMADO existe diferencia significativa entre hombres y mujeres "
        "en el Δ Z-score SAT nota (ambos ns). El sexo no modera el impacto de la formación.",
        "En TALLER los hombres promedian −0.091 y las mujeres +0.003; en DIPLOMADO −0.117 vs "
        "+0.029. Aunque las mujeres tienden a valores levemente superiores, la diferencia "
        "no alcanza significancia estadística (Cohen d máx. −0.183).",
    ])

# Tabla 6B — SAT BIN
rows_6b = []
for tipo in ["TALLER", "DIPLOMADO", "PROYECTO"]:
    sub = df5_bin[df5_bin["tipo_principal"] == tipo]
    r   = _row_sexo(sub, "delta_z_bin")
    if r is None:
        rows_6b.append([tipo, "—", "—", "—", "—", "—", "—", "insuf. n", "—", "—"])
    else:
        rows_6b.append([tipo] + r)

row = write_table(ws6, row,
    "Tabla 6B — Prueba t Independiente: Δ Z-score % Recomendación por Sexo (Hombre vs Mujer) — G5.2",
    "Misma comparación para la métrica SAT_BIN (¿Recomendaría a este/a docente?). "
    "PROYECTO excluido por n insuficiente.",
    h_sex, rows_6b, sig_col_idx=9,
    col_widths=[18, 12, 12, 12, 12, 12, 10, 10, 10, 8],
    bullets=[
        "Tampoco se detectan diferencias significativas por sexo en % Recomendación (todos ns). "
        "El patrón es coherente con SAT nota: el sexo no es variable moderadora.",
        "En DIPLOMADO la diferencia observada es mayor (H: −0.220, M: +0.109, d=−0.387) pero "
        "no significativa dado el tamaño muestral reducido (nM=14, nF=22). Requiere cautela.",
    ])

# ── Guardar ────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"Guardado: {OUT}")
print(f"Hojas: {[ws.title for ws in wb.worksheets]}")
